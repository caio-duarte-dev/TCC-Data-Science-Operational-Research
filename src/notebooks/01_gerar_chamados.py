import pandas as pd
from openpyxl import load_workbook
import os
import random
from datetime import timedelta
from pathlib import Path

# ==============================
# Definição de caminhos
# ==============================

BASE_DIR = Path(__file__).resolve().parents[1]
DATA_DIR = BASE_DIR / "data"

arquivo = DATA_DIR / "dados.xlsx"
path_chamados = DATA_DIR / "chamados.xlsx"

# ==============================
# Leitura dos dados base
# ==============================

abas = pd.read_excel(arquivo, sheet_name=None)

clientes = abas.get('clientes')
maquinas = abas.get('maquinas')
tipo_manutencao = abas.get('tipo_manutencao')
prioridades = abas.get('prioridades')
datas = abas.get('datas')

# ==============================
# Função para append no Excel
# ==============================

def append_chamados(novos_chamados, path_excel):
    try:
        book = load_workbook(path_excel)
        sheet = book.active
        max_row = sheet.max_row

        with pd.ExcelWriter(path_excel, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            novos_chamados.to_excel(writer, index=False, header=False, startrow=max_row)

    except FileNotFoundError:
        novos_chamados.to_excel(path_excel, index=False, header=True)

# ==============================
# Simulação de chamados
# ==============================

def simular_chamados(qtd, clientes, maquinas, tipo_manutencao, prioridades, datas, path_excel):

    if os.path.exists(path_excel):
        chamados_existentes = pd.read_excel(path_excel)
        ultimo_id = chamados_existentes['id'].max()
        if pd.isna(ultimo_id):
            ultimo_id = 0
    else:
        colunas = ['id', 'cliente', 'maquina', 'tipo_manutencao', 'tempo_necessario', 'prioridade', 'data_abertura']
        df_vazio = pd.DataFrame(columns=colunas)
        df_vazio.to_excel(path_excel, index=False)
        ultimo_id = 0

    chamados = []

    data_inicio = pd.to_datetime(datas['data_inicio'].iloc[0])
    data_fim = pd.to_datetime(datas['data_fim'].iloc[0])
    delta_dias = (data_fim - data_inicio).days

    for i in range(1, qtd + 1):
        cliente = random.choice(clientes['clientes'].tolist())
        maquina = random.choice(maquinas['maquinas'].tolist())

        tipo_row = tipo_manutencao.sample(1).iloc[0]
        tipo = tipo_row['tipo_manutencao']
        tempo = tipo_row['tempo_necessario']

        prioridade = random.choice(prioridades['prioridade'].tolist())

        data_abertura = (
            data_inicio + timedelta(days=random.randint(0, delta_dias))
        ).strftime('%Y-%m-%d')

        chamados.append({
            'id': ultimo_id + i,
            'cliente': cliente,
            'maquina': maquina,
            'tipo_manutencao': tipo,
            'tempo_necessario': tempo,
            'prioridade': prioridade,
            'data_abertura': data_abertura
        })

    novos_chamados = pd.DataFrame(chamados)

    append_chamados(novos_chamados, path_excel)

    return novos_chamados

# ==============================
# Execução
# ==============================

novos_chamados = simular_chamados(
    qtd=200,
    clientes=clientes,
    maquinas=maquinas,
    tipo_manutencao=tipo_manutencao,
    prioridades=prioridades,
    datas=datas,
    path_excel=path_chamados
)

chamados = pd.read_excel(path_chamados)

print(chamados)

